# coding:utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time
import os
class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None)
        self.RGBDict = {'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathandName):
        try:
            self.workbook = openpyxl.load_workbook(excelPathandName)
        except Exception as e:
            raise e
        self.excelFile = excelPathandName
        return self.workbook

    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        return sheet.max_row

    def getColsNumber(self,sheet):
        return sheet.max_column

    def getStartColNumber(self,sheet):
        return sheet.min_column

    def getStartRowNumber(self,sheet):
        return sheet.min_row

    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回所有数据元组，下标从1开始
        try:
            return sheet.rows[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        #获取sheet中某一列，返回所有数据元组，下标从1开始
        try:
            return list(sheet.columns)[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,rowNo = None,colsNo = None,coordinate = None):
        #coordinate取如‘A1' ,行列指定到具体位置
        if coordinate != None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self,sheet,rowNo = None,colsNo = None,coordinate = None):
        if coordinate != None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCell(self,sheet,content,rowNo = None,colsNo = None,coordinate = None,style = None):
        if coordinate is not None:
            try:
                sheet[coordinate].value = content
                if style is not None:
                    sheet[coordinate].font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCellCurrentTime(self,sheet,rowNo = None,colsNo = None,coordinate = None):
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet[coordinate].value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

if __name__ == '__main__':
    data_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excelPath = os.path.join(data_path, "apitest.xlsx")
    vv=ParseExcel()
    vv.loadWorkBook(excelPath)
    b=vv.getSheetByIndex(0)
    a = vv.getSheetByName('jktest')
    print (vv.getSheetByName('jktest').title)
    print ((vv.getRowsNumber(a))+1)
    cc = vv.getColumn(b,2)
    z = vv.getCellOfObject(b,coordinate = 'A12')
    # for i in cc:
    #     print (i.value)
    y = eval(vv.getCellOfValue(b,2,7))  #eval 类型转换
    print (y)
    # print (y['status'])
    print (type(y))
    print (vv.getCellOfValue(b,3,7))
    vv.writeCell(b,"fail",3,7,style = 'red')
    # print (vv.writeCell(b,"fail",3,7,style = 'red'))
    # print (vv.writeCellCurrentTime(b,3,8))
    # print (vv.writeCell(b,"fail",coordinate = 'A9',style = 'red'))
    # print (vv.writeCellCurrentTime(b,coordinate = 'A10'))
    # print (vv.getRowsNumber(a))


